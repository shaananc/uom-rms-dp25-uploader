from __future__ import annotations

import argparse
import json
import logging
from pathlib import Path

import yaml
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


LOGGER = logging.getLogger(__name__)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export RMS budget payload JSON from the spreadsheet."
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("config.yml"),
        help="Path to the ARC budget config file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("tmp/rms_current_payload.json"),
        help="Path to write the generated RMS payload JSON.",
    )
    return parser.parse_args()


def money(value: int | float | None) -> int:
    if value in (None, ""):
        return 0
    return int(round(float(value)))


def first_line(value: str | None) -> str:
    return (value or "").splitlines()[0].strip()


def offset_cells(row: int, year_columns: list[str], offset: int) -> list[str]:
    cells: list[str] = []
    for column in year_columns:
        start_idx = column_index_from_string(column)
        cells.append(f"{get_column_letter(start_idx + offset)}{row}")
    return cells


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    args = parse_args()

    config = yaml.safe_load(args.config.read_text())
    workbook = load_workbook(config["xlsx_path"], keep_links=False)
    sheet = workbook[config["sheet_name"]]
    year_columns = config["year_columns"]

    years = [{"year": index + 1, "entries": []} for index in range(len(year_columns))]

    for row_config in config["row_configs"]:
        category = row_config["category"].title()
        for row in range(row_config["start_row"], row_config["end_row"] + 1):
            name = first_line(sheet[f"A{row}"].value)
            if not name:
                continue

            arc_values = [money(sheet[cell].value) for cell in offset_cells(row, year_columns, 0)]
            admin_values = [money(sheet[cell].value) for cell in offset_cells(row, year_columns, 1)]
            inkind_values = [money(sheet[cell].value) for cell in offset_cells(row, year_columns, 2)]

            for year_index, year_data in enumerate(years):
                arc = arc_values[year_index]
                admin = admin_values[year_index]
                inkind = inkind_values[year_index]
                if not any((arc, admin, inkind)):
                    continue
                year_data["entries"].append(
                    {
                        "category": category,
                        "name": name,
                        "arc": arc,
                        "admin": admin,
                        "inkind": inkind,
                    }
                )

    payload = {
        "proposal_id": str(config["proposal_id"]),
        "years": years,
    }
    args.output.write_text(json.dumps(payload, indent=2) + "\n")
    LOGGER.info("Wrote %s from %s", args.output, config["xlsx_path"])


if __name__ == "__main__":
    main()
