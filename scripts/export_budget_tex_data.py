from __future__ import annotations

import argparse
import logging
import re
from pathlib import Path

from rich.logging import RichHandler
import yaml
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


LOGGER = logging.getLogger(__name__)
YEAR_WORDS = ["One", "Two", "Three", "Four", "Five"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export DP27 budget figures from the spreadsheet into TeX macros."
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("config.yml"),
        help="Path to the ARC budget config file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(
            "/Users/shaananc/Library/CloudStorage/OneDrive-Personal/grants/arc-dp-2027/budget-justification/spreadsheet-budget.tex"
        ),
        help="Path to write the generated TeX macro file.",
    )
    return parser.parse_args()


def money(value: int | float | None) -> int:
    if value in (None, ""):
        return 0
    return int(round(float(value)))


def money_text(value: int) -> str:
    return f"{value:,}"


def first_line(value: str | None) -> str:
    return (value or "").splitlines()[0].strip()


def escape_tex(value: str) -> str:
    replacements = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    return "".join(replacements.get(char, char) for char in value)


def parse_breakdown(text: str) -> tuple[int, int, int, int]:
    amounts = [money(raw.replace(",", "")) for raw in re.findall(r"\$([\d,]+)", text)]
    if len(amounts) < 4:
        raise ValueError(f"Could not parse four travel components from: {text!r}")
    return tuple(amounts[:4])


def arc_cells(row: int, year_columns: list[str]) -> list[str]:
    return [f"{column}{row}" for column in year_columns]


def admin_cells(row: int, year_columns: list[str]) -> list[str]:
    cells: list[str] = []
    for column in year_columns:
        start_idx = column_index_from_string(column)
        cells.append(f"{get_column_letter(start_idx + 1)}{row}")
    return cells


def inkind_cells(row: int, year_columns: list[str]) -> list[str]:
    cells: list[str] = []
    for column in year_columns:
        start_idx = column_index_from_string(column)
        cells.append(f"{get_column_letter(start_idx + 2)}{row}")
    return cells


def find_named_row(rows: list[dict[str, object]], needle: str) -> dict[str, object]:
    for row in rows:
        if needle in str(row["name"]):
            return row
    raise ValueError(f"Could not find row matching {needle!r}")


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(message)s", handlers=[RichHandler()])
    args = parse_args()

    config = yaml.safe_load(args.config.read_text())
    workbook_path = Path(config["xlsx_path"])
    sheet_name = config["sheet_name"]
    year_columns = config["year_columns"]

    workbook = load_workbook(workbook_path, keep_links=False)
    sheet = workbook[sheet_name]

    categories = {item["category"]: item for item in config["row_configs"]}

    personnel_rows: list[dict[str, object]] = []
    for row in range(
        categories["PERSONNEL"]["start_row"], categories["PERSONNEL"]["end_row"] + 1
    ):
        name = sheet[f"A{row}"].value
        if not name:
            continue
        personnel_rows.append(
            {
                "name": first_line(str(name)),
                "arc": [money(sheet[cell].value) for cell in arc_cells(row, year_columns)],
                "inkind": [
                    money(sheet[cell].value) for cell in inkind_cells(row, year_columns)
                ],
            }
        )

    ci_yuan = find_named_row(personnel_rows, "CI Xingliang Yuan")
    ci_cohney = find_named_row(personnel_rows, "CI Shaanan Cohney")
    ci_lai = find_named_row(personnel_rows, "CI Shangqi Lai")
    hdr_topup = find_named_row(personnel_rows, "HDR (Higher Degree by Research stipend)")
    hdr_arc = find_named_row(personnel_rows, "ARC funded")
    pdra = find_named_row(personnel_rows, "Postdoctoral Research Associate")

    travel_rows: list[dict[str, object]] = []
    for row in range(categories["TRAVEL"]["start_row"], categories["TRAVEL"]["end_row"] + 1):
        name = sheet[f"A{row}"].value
        if not name:
            continue
        arc_values = [money(sheet[cell].value) for cell in arc_cells(row, year_columns)]
        total = sum(arc_values)
        if not total:
            continue
        year_index = next(index for index, value in enumerate(arc_values) if value)
        airfare, accommodation, registration, local = parse_breakdown(str(name))
        travel_rows.append(
            {
                "title": first_line(str(name)),
                "year_index": year_index,
                "total": total,
                "airfare": airfare,
                "accommodation": accommodation,
                "registration": registration,
                "local": local,
            }
        )
    travel_rows.sort(key=lambda item: int(item["year_index"]))

    other_arc_values: list[int] = []
    other_rows: list[dict[str, object]] = []
    for row in range(categories["OTHER"]["start_row"], categories["OTHER"]["end_row"] + 1):
        name = sheet[f"A{row}"].value
        if not name:
            continue
        arc_values = [money(sheet[cell].value) for cell in arc_cells(row, year_columns)]
        other_rows.append(
            {
                "name": first_line(str(name)),
                "arc": arc_values,
                "admin": [money(sheet[cell].value) for cell in admin_cells(row, year_columns)],
                "inkind": [
                    money(sheet[cell].value) for cell in inkind_cells(row, year_columns)
                ],
            }
        )
        other_arc_values.extend(arc_values)

    publication_row = find_named_row(other_rows, "Publication Fees for Gold Open Access")
    workshop_row = find_named_row(other_rows, "Industry dissemination workshop")
    postdoc_laptop_row = find_named_row(other_rows, "Postdoctoral researcher laptop")
    hdr_laptop_row = find_named_row(other_rows, "HDR candidate laptop")
    hdr_travel_support_row = find_named_row(other_rows, "FEIT HDR conference-travel support")

    personnel_arc_total = sum(pdra["arc"]) + sum(hdr_arc["arc"])
    travel_arc_total = sum(int(row["total"]) for row in travel_rows)
    other_arc_total = sum(other_arc_values)
    arc_total = personnel_arc_total + travel_arc_total + other_arc_total

    ci_yuan_total = sum(ci_yuan["inkind"])
    ci_cohney_total = sum(ci_cohney["inkind"])
    ci_lai_total = sum(ci_lai["inkind"])
    hdr_topup_total = sum(hdr_topup["inkind"])
    personnel_inkind_total = (
        ci_yuan_total + ci_cohney_total + ci_lai_total + hdr_topup_total
    )
    other_inkind_total = sum(sum(row["inkind"]) for row in other_rows)
    admin_cash_total = sum(sum(row["admin"]) for row in other_rows)
    inkind_total = personnel_inkind_total + other_inkind_total
    non_arc_total = inkind_total + admin_cash_total
    other_arc_years = [sum(row["arc"][i] for row in other_rows) for i in range(len(year_columns))]
    other_non_arc_years = [
        sum(row["admin"][i] + row["inkind"][i] for row in other_rows)
        for i in range(len(year_columns))
    ]
    other_non_arc_total = sum(other_non_arc_years)

    lines = [
        "% Auto-generated from the DP27 budget spreadsheet.",
        "% Regenerate with: uv run python scripts/export_budget_tex_data.py",
    ]

    macros = {
        "BudgetArcTotal": money_text(arc_total),
        "BudgetPersonnelArcTotal": money_text(personnel_arc_total),
        "BudgetTravelArcTotal": money_text(travel_arc_total),
        "BudgetOtherArcTotal": money_text(other_arc_total),
        "BudgetOtherArcYearOne": money_text(other_arc_years[0]),
        "BudgetOtherArcYearTwo": money_text(other_arc_years[1]),
        "BudgetOtherArcYearThree": money_text(other_arc_years[2]),
        "BudgetPdraYearOne": money_text(pdra["arc"][0]),
        "BudgetPdraYearTwo": money_text(pdra["arc"][1]),
        "BudgetPdraYearThree": money_text(pdra["arc"][2]),
        "BudgetPdraTotal": money_text(sum(pdra["arc"])),
        "BudgetHdrArcYearOne": money_text(hdr_arc["arc"][0]),
        "BudgetHdrArcYearTwo": money_text(hdr_arc["arc"][1]),
        "BudgetHdrArcYearThree": money_text(hdr_arc["arc"][2]),
        "BudgetHdrArcTotal": money_text(sum(hdr_arc["arc"])),
        "BudgetCiYuanYearOne": money_text(ci_yuan["inkind"][0]),
        "BudgetCiYuanYearTwo": money_text(ci_yuan["inkind"][1]),
        "BudgetCiYuanYearThree": money_text(ci_yuan["inkind"][2]),
        "BudgetCiYuanTotal": money_text(ci_yuan_total),
        "BudgetCiCohneyYearOne": money_text(ci_cohney["inkind"][0]),
        "BudgetCiCohneyYearTwo": money_text(ci_cohney["inkind"][1]),
        "BudgetCiCohneyYearThree": money_text(ci_cohney["inkind"][2]),
        "BudgetCiCohneyTotal": money_text(ci_cohney_total),
        "BudgetCiLaiYearOne": money_text(ci_lai["inkind"][0]),
        "BudgetCiLaiYearTwo": money_text(ci_lai["inkind"][1]),
        "BudgetCiLaiYearThree": money_text(ci_lai["inkind"][2]),
        "BudgetCiLaiTotal": money_text(ci_lai_total),
        "BudgetHdrTopupYearOne": money_text(hdr_topup["inkind"][0]),
        "BudgetHdrTopupYearTwo": money_text(hdr_topup["inkind"][1]),
        "BudgetHdrTopupYearThree": money_text(hdr_topup["inkind"][2]),
        "BudgetHdrTopupTotal": money_text(hdr_topup_total),
        "BudgetPersonnelInKindTotal": money_text(personnel_inkind_total),
        "BudgetAdminCashTotal": money_text(admin_cash_total),
        "BudgetInKindTotal": money_text(inkind_total),
        "BudgetNonArcTotal": money_text(non_arc_total),
        "BudgetOtherNonArcYearOne": money_text(other_non_arc_years[0]),
        "BudgetOtherNonArcYearTwo": money_text(other_non_arc_years[1]),
        "BudgetOtherNonArcYearThree": money_text(other_non_arc_years[2]),
        "BudgetOtherNonArcTotal": money_text(other_non_arc_total),
        "BudgetPublicationYearOne": money_text(publication_row["arc"][0]),
        "BudgetPublicationYearTwo": money_text(publication_row["arc"][1]),
        "BudgetPublicationYearThree": money_text(publication_row["arc"][2]),
        "BudgetPublicationTotal": money_text(sum(publication_row["arc"])),
        "BudgetWorkshopYearOne": money_text(workshop_row["arc"][0]),
        "BudgetWorkshopYearTwo": money_text(workshop_row["arc"][1]),
        "BudgetWorkshopYearThree": money_text(workshop_row["arc"][2]),
        "BudgetWorkshopTotal": money_text(sum(workshop_row["arc"])),
        "BudgetPostdocLaptopYearOne": money_text(postdoc_laptop_row["inkind"][0]),
        "BudgetPostdocLaptopYearTwo": money_text(postdoc_laptop_row["inkind"][1]),
        "BudgetPostdocLaptopYearThree": money_text(postdoc_laptop_row["inkind"][2]),
        "BudgetPostdocLaptopTotal": money_text(sum(postdoc_laptop_row["inkind"])),
        "BudgetHdrLaptopYearOne": money_text(hdr_laptop_row["inkind"][0]),
        "BudgetHdrLaptopYearTwo": money_text(hdr_laptop_row["inkind"][1]),
        "BudgetHdrLaptopYearThree": money_text(hdr_laptop_row["inkind"][2]),
        "BudgetHdrLaptopTotal": money_text(sum(hdr_laptop_row["inkind"])),
        "BudgetEquipmentYearOne": money_text(
            postdoc_laptop_row["inkind"][0] + hdr_laptop_row["inkind"][0]
        ),
        "BudgetEquipmentYearTwo": money_text(
            postdoc_laptop_row["inkind"][1] + hdr_laptop_row["inkind"][1]
        ),
        "BudgetEquipmentYearThree": money_text(
            postdoc_laptop_row["inkind"][2] + hdr_laptop_row["inkind"][2]
        ),
        "BudgetEquipmentTotal": money_text(
            sum(postdoc_laptop_row["inkind"]) + sum(hdr_laptop_row["inkind"])
        ),
        "BudgetHdrTravelSupportYearOne": money_text(hdr_travel_support_row["admin"][0]),
        "BudgetHdrTravelSupportYearTwo": money_text(hdr_travel_support_row["admin"][1]),
        "BudgetHdrTravelSupportYearThree": money_text(hdr_travel_support_row["admin"][2]),
        "BudgetHdrTravelSupportTotal": money_text(sum(hdr_travel_support_row["admin"])),
    }

    for index, row in enumerate(travel_rows):
        year_word = YEAR_WORDS[index]
        macros[f"BudgetTravelYear{year_word}Title"] = escape_tex(str(row["title"]))
        macros[f"BudgetTravelYear{year_word}Total"] = money_text(int(row["total"]))
        macros[f"BudgetTravelYear{year_word}Airfare"] = money_text(int(row["airfare"]))
        macros[f"BudgetTravelYear{year_word}Accommodation"] = money_text(
            int(row["accommodation"])
        )
        macros[f"BudgetTravelYear{year_word}Registration"] = money_text(
            int(row["registration"])
        )
        macros[f"BudgetTravelYear{year_word}Local"] = money_text(int(row["local"]))

    for name, value in macros.items():
        lines.append(rf"\providecommand{{\{name}}}{{{value}}}")

    args.output.write_text("\n".join(lines) + "\n")
    LOGGER.info("Wrote %s from %s", args.output, workbook_path)


if __name__ == "__main__":
    main()
